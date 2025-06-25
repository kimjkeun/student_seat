import pandas as pd
import openpyxl
import re

try:
    # Excel 파일 열기
    excel_file = 'teacher_view_seating_chart.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    
    print("=== 자리표 Excel 파일 분석 ===")
    print(f"파일명: {excel_file}")
    print(f"시트 수: {len(wb.sheetnames)}개")
    print(f"시트 목록: {wb.sheetnames}")
    print()
    
    all_seating_data = []
    
    # 각 시트 분석
    for sheet_name in wb.sheetnames:
        print(f"=== {sheet_name} 분석 ===")
        
        # pandas로 시트 읽기
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
        print(f"크기: {df.shape[0]}행 x {df.shape[1]}열")
        
        # 클래스 번호 추출 (3-5(C) -> 5)
        class_match = re.search(r'3-(\d+)', sheet_name)
        class_num = class_match.group(1) if class_match else "unknown"
        
        students_in_sheet = []
        
        # 각 셀에서 학번과 이름 추출
        for row_idx, row in df.iterrows():
            for col_idx, cell_value in enumerate(row):
                if pd.notna(cell_value) and isinstance(cell_value, str):
                    cell_text = cell_value.strip()
                    
                    # 학번 이름 패턴 찾기 (예: "30102 권민지")
                    student_pattern = re.search(r'(\d{5})\s+([가-힣]{2,4})', cell_text)
                    if student_pattern:
                        student_id, student_name = student_pattern.groups()
                        
                        student_info = {
                            '시트': sheet_name,
                            '클래스': class_num,
                            '학번': student_id,
                            '이름': student_name,
                            '행': row_idx + 1,
                            '열': col_idx + 1,
                            '좌석위치': f'({row_idx + 1}, {col_idx + 1})'
                        }
                        students_in_sheet.append(student_info)
                        all_seating_data.append(student_info)
        
        print(f"발견된 학생: {len(students_in_sheet)}명")
        
        # 자리 배치 시각화
        print(f"\n{sheet_name} 자리 배치:")
        seat_layout = []
        for row_idx in range(df.shape[0]):
            row_seats = []
            for col_idx in range(df.shape[1]):
                cell_value = df.iloc[row_idx, col_idx]
                if pd.notna(cell_value) and isinstance(cell_value, str):
                    cell_text = cell_value.strip()
                    # 학번 이름 패턴이 있으면 이름만 표시
                    student_match = re.search(r'\d{5}\s+([가-힣]{2,4})', cell_text)
                    if student_match:
                        row_seats.append(student_match.group(1))
                    elif cell_text == "교탁":
                        row_seats.append("교탁")
                    else:
                        row_seats.append("")
                else:
                    row_seats.append("")
            seat_layout.append(row_seats)
        
        # 자리 배치 출력 (교사 시점)
        for row in seat_layout:
            formatted_row = []
            for seat in row:
                if seat:
                    formatted_row.append(f"{seat:^8}")
                else:
                    formatted_row.append("        ")
            print("  " + " | ".join(formatted_row))
        
        print("\n" + "="*60 + "\n")
    
    # 전체 결과 정리
    print("=== 전체 자리표 분석 결과 ===")
    print(f"총 학생 수: {len(all_seating_data)}명")
    
    if all_seating_data:
        # DataFrame으로 변환
        df_seating = pd.DataFrame(all_seating_data)
        
        # 클래스별 학생 수
        print("\n클래스별 학생 수:")
        class_counts = df_seating['클래스'].value_counts().sort_index()
        for class_num, count in class_counts.items():
            print(f"  클래스 {class_num}: {count}명")
        
        # CSV로 저장
        df_seating.to_csv('seating_chart_analysis.csv', index=False, encoding='utf-8-sig')
        print(f"\n자리표 분석 결과가 'seating_chart_analysis.csv'에 저장되었습니다.")
        
        # 샘플 데이터 출력
        print("\n샘플 자리표 데이터:")
        print(df_seating.head(10).to_string(index=False))
        
        # 각 클래스별 상세 정보
        print("\n=== 클래스별 상세 정보 ===")
        for class_num in sorted(df_seating['클래스'].unique()):
            class_students = df_seating[df_seating['클래스'] == class_num]
            print(f"\n클래스 {class_num} ({len(class_students)}명):")
            for _, student in class_students.iterrows():
                print(f"  {student['학번']} {student['이름']} - 좌석: {student['좌석위치']}")

except Exception as e:
    print(f"파일 분석 오류: {e}")
    import traceback
    traceback.print_exc()
