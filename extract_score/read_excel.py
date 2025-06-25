import openpyxl
import sys

try:
    # 엑셀 파일 열기
    wb = openpyxl.load_workbook('score.xlsx')
    print("엑셀 파일을 성공적으로 열었습니다!")
    
    # 워크시트 정보
    print(f"워크시트명: {wb.sheetnames}")
    
    # 활성 시트 선택
    ws = wb.active
    print(f"활성 시트: {ws.title}")
    print(f"최대 행: {ws.max_row}")
    print(f"최대 열: {ws.max_column}")
    
    print("\n=== 데이터 내용 ===")
    
    # 모든 데이터 출력 (행별로)
    for row in range(1, ws.max_row + 1):
        row_data = []
        has_data = False
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                has_data = True
            row_data.append(cell_value)
        
        # 데이터가 있는 행만 출력
        if has_data:
            print(f"행 {row:2d}: ", end="")
            for i, value in enumerate(row_data):
                if value is not None:
                    if isinstance(value, float):
                        print(f"[{i+1}]{value:.1f}", end=" ")
                    else:
                        print(f"[{i+1}]{value}", end=" ")
            print()  # 줄바꿈

except ImportError:
    print("openpyxl 라이브러리가 설치되어 있지 않습니다.")
    print("다음 명령어로 설치하세요: pip install openpyxl")
except FileNotFoundError:
    print("score.xlsx 파일을 찾을 수 없습니다.")
except Exception as e:
    print(f"오류가 발생했습니다: {e}")
