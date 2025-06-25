import pandas as pd
import openpyxl

# openpyxl로 먼저 구조 파악
wb = openpyxl.load_workbook('score.xlsx')
ws = wb.active

print("=== 엑셀 파일 구조 분석 ===")
print(f"워크시트: {ws.title}")
print(f"크기: {ws.max_row}행 x {ws.max_column}열")
print()

print("=== 헤더 정보 (처음 5행) ===")
for row in range(1, 6):
    row_data = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        if value is not None:
            row_data.append(f"[{col}]{value}")
    if row_data:
        print(f"행 {row}: {' '.join(row_data)}")

print("\n=== 데이터 부분 추정 ===")
# 숫자가 많이 포함된 행들을 찾아서 데이터 영역 추정
data_rows = []
for row in range(6, ws.max_row + 1):
    numeric_count = 0
    row_values = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        if value is not None:
            row_values.append(value)
            if isinstance(value, (int, float)):
                numeric_count += 1
    
    if numeric_count >= 3:  # 숫자가 3개 이상인 행을 데이터로 간주
        data_rows.append((row, row_values, numeric_count))

print(f"데이터로 추정되는 행: {len(data_rows)}개")
print("\n=== 학생 성적 데이터 (추정) ===")
for i, (row_num, values, num_count) in enumerate(data_rows[:10]):  # 처음 10개만 표시
    print(f"행 {row_num:2d} (숫자 {num_count}개): {values}")

if len(data_rows) > 10:
    print(f"... (총 {len(data_rows)}개 행 중 처음 10개만 표시)")

# pandas로 데이터 부분만 읽기 시도
if data_rows:
    start_row = data_rows[0][0] - 1  # 0-based index
    print(f"\n=== 데이터 영역 상세 분석 (행 {start_row+1}부터) ===")
    
    try:
        # 데이터 영역만 읽기
        df = pd.read_excel('score.xlsx', skiprows=start_row-1, header=None)
        df_clean = df.dropna(how='all').dropna(axis=1, how='all')
        
        print(f"정리된 데이터 크기: {df_clean.shape[0]}행 x {df_clean.shape[1]}열")
        print("\n처음 10행:")
        print(df_clean.head(10).to_string())
        
        print("\n숫자 컬럼 통계:")
        numeric_df = df_clean.select_dtypes(include=['number'])
        if not numeric_df.empty:
            print(numeric_df.describe())
        
    except Exception as e:
        print(f"pandas 분석 중 오류: {e}")

print("\n=== 푸터 정보 (마지막 5행) ===")
for row in range(max(1, ws.max_row-4), ws.max_row + 1):
    row_data = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        if value is not None:
            row_data.append(f"[{col}]{value}")
    if row_data:
        print(f"행 {row}: {' '.join(row_data)}")
