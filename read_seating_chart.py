import pandas as pd
import openpyxl

try:
    # Excel 파일 열기
    excel_file = 'teacher_view_seating_chart.xlsx'
    
    # openpyxl로 시트 정보 확인
    wb = openpyxl.load_workbook(excel_file)
    print("=== Excel 파일 정보 ===")
    print(f"파일명: {excel_file}")
    print(f"시트 수: {len(wb.sheetnames)}개")
    print(f"시트 목록: {wb.sheetnames}")
    print()
    
    # 각 시트 분석
    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"=== 시트 {sheet_idx}: {sheet_name} ===")
        
        # pandas로 시트 읽기
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
            print(f"크기: {df.shape[0]}행 x {df.shape[1]}열")
            
            # 데이터 미리보기
            print("\n데이터 미리보기:")
            print(df.head(10).to_string(index=False, header=False, na_rep=''))
            
            # 비어있지 않은 셀 개수 확인
            non_empty_cells = df.count().sum()
            print(f"\n비어있지 않은 셀 개수: {non_empty_cells}개")
            
            # 학생 이름으로 보이는 패턴 찾기 (한글 이름)
            import re
            student_names = []
            for row_idx, row in df.iterrows():
                for col_idx, cell_value in enumerate(row):
                    if pd.notna(cell_value) and isinstance(cell_value, str):
                        # 한글 이름 패턴 (2-4글자 한글)
                        if re.match(r'^[가-힣]{2,4}$', cell_value.strip()):
                            student_names.append({
                                '이름': cell_value.strip(),
                                '행': row_idx + 1,
                                '열': col_idx + 1,
                                '위치': f'({row_idx + 1}, {col_idx + 1})'
                            })
            
            if student_names:
                print(f"\n발견된 학생 이름: {len(student_names)}명")
                print("학생 이름 목록:")
                for i, student in enumerate(student_names[:10], 1):  # 처음 10명만 표시
                    print(f"  {i:2d}. {student['이름']} - 위치: {student['위치']}")
                if len(student_names) > 10:
                    print(f"  ... 외 {len(student_names) - 10}명")
                
                # 자리 배치 시각화 (간단한 형태)
                print(f"\n=== {sheet_name} 자리 배치도 ===")
                seat_chart = df.copy()
                # 학생 이름만 표시하고 나머지는 공백으로
                for row_idx in range(seat_chart.shape[0]):
                    for col_idx in range(seat_chart.shape[1]):
                        cell_value = seat_chart.iloc[row_idx, col_idx]
                        if pd.notna(cell_value) and isinstance(cell_value, str):
                            if not re.match(r'^[가-힣]{2,4}$', cell_value.strip()):
                                seat_chart.iloc[row_idx, col_idx] = ''
                        else:
                            seat_chart.iloc[row_idx, col_idx] = ''
                
                print(seat_chart.to_string(index=False, header=False, na_rep=''))
            else:
                print("\n학생 이름을 찾을 수 없습니다.")
            
        except Exception as e:
            print(f"시트 읽기 오류: {e}")
        
        print("\n" + "="*50 + "\n")
    
    # 전체 시트의 학생 정보를 하나로 합치기
    print("=== 전체 시트 통합 분석 ===")
    all_students = []
    
    for sheet_name in wb.sheetnames:
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
            
            for row_idx, row in df.iterrows():
                for col_idx, cell_value in enumerate(row):
                    if pd.notna(cell_value) and isinstance(cell_value, str):
                        if re.match(r'^[가-힣]{2,4}$', cell_value.strip()):
                            all_students.append({
                                '시트': sheet_name,
                                '이름': cell_value.strip(),
                                '행': row_idx + 1,
                                '열': col_idx + 1
                            })
        except:
            continue
    
    if all_students:
        print(f"전체 학생 수: {len(all_students)}명")
        
        # 시트별 학생 수
        import pandas as pd
        df_students = pd.DataFrame(all_students)
        sheet_counts = df_students['시트'].value_counts()
        print("\n시트별 학생 수:")
        for sheet, count in sheet_counts.items():
            print(f"  {sheet}: {count}명")
        
        # CSV로 저장
        df_students.to_csv('seating_chart_students.csv', index=False, encoding='utf-8-sig')
        print(f"\n자리표 학생 정보가 'seating_chart_students.csv'에 저장되었습니다.")

except Exception as e:
    print(f"파일 읽기 오류: {e}")
    import traceback
    traceback.print_exc()
